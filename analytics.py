import os
import psycopg2
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import numpy as np

# создаём папки
os.makedirs("charts", exist_ok=True)
os.makedirs("exports", exist_ok=True)

# ------------------------------
# Подключение к БД
# ------------------------------
conn = psycopg2.connect(
    dbname="MAXIMVISUAL",
    user="postgres",
    password="maxim22s2",
    host="localhost",
    port="5433"
)

# ------------------------------
# SQL-запросы
# ------------------------------
queries = {
    "orders_by_status": """
        SELECT order_status, COUNT(*) AS total
        FROM orders
        GROUP BY order_status;
    """,
    "orders_by_state": """
        SELECT customer_state, COUNT(o.order_id) AS total
        FROM customers c
        JOIN orders o ON c.customer_id = o.customer_id
        GROUP BY customer_state
        ORDER BY total DESC
        LIMIT 10;
    """,
    "avg_freight_by_state": """
        SELECT customer_state, ROUND(AVG(freight_value),2) AS avg_freight
        FROM order_items oi
        JOIN orders o ON oi.order_id = o.order_id
        JOIN customers c ON o.customer_id = c.customer_id
        GROUP BY customer_state
        ORDER BY avg_freight DESC
        LIMIT 10;
    """,
    "orders_by_month": """
        SELECT DATE_TRUNC('month', order_purchase_timestamp) AS month, COUNT(*) AS total
        FROM orders
        GROUP BY month
        ORDER BY month;
    """,
    "order_values": """
        SELECT SUM(price) AS order_value
        FROM order_items
        GROUP BY order_id;
    """,
    "order_items_vs_value": """
        SELECT COUNT(*) AS items_count, SUM(price) AS order_value
        FROM order_items
        GROUP BY order_id;
    """
}

dfs = {name: pd.read_sql(query, conn) for name, query in queries.items()}

# ------------------------------
# 1. Статичные графики
# ------------------------------
sns.set(style="whitegrid", font_scale=1.1)

# 1. Pie chart (Orders by Status)
plt.figure(figsize=(8,8))
sizes = dfs["orders_by_status"]["total"]
labels = dfs["orders_by_status"]["order_status"]

sorted_data = sorted(zip(labels, sizes), key=lambda x: x[1], reverse=True)
labels, sizes = zip(*sorted_data)

wedges, texts = plt.pie(
    sizes,
    startangle=140,
    autopct=None,
    pctdistance=0.85
)

centre_circle = plt.Circle((0,0),0.70,fc='white')
fig = plt.gcf()
fig.gca().add_artist(centre_circle)

total = sum(sizes)
legend_labels = [
    f"{label} – {size} ({size/total*100:.1f}%)"
    for label, size in zip(labels, sizes)
]

plt.legend(
    wedges,
    legend_labels,
    title="Order Status",
    loc="center left",
    bbox_to_anchor=(1, 0, 0.5, 1)
)

plt.title("Orders by Status", fontsize=16)
plt.tight_layout()
plt.savefig("charts/orders_by_status_pie.png")
plt.clf()

# 2. Bar chart (Top States by Orders)
plt.figure(figsize=(10,6))
sns.barplot(data=dfs["orders_by_state"], x="customer_state", y="total", palette="Blues_d")
plt.title("Top 10 States by Orders", fontsize=16)
plt.xlabel("State")
plt.ylabel("Orders")
plt.tight_layout()
plt.savefig("charts/orders_by_state_bar.png")
plt.clf()

# 3. Horizontal Bar (Avg Freight by State)
plt.figure(figsize=(10,6))
sns.barplot(data=dfs["avg_freight_by_state"], y="customer_state", x="avg_freight", palette="Greens_d")
plt.title("Avg Freight by State", fontsize=16)
plt.xlabel("Avg Freight Value")
plt.ylabel("State")
plt.tight_layout()
plt.savefig("charts/avg_freight_by_state_hbar.png")
plt.clf()

# 4. Line chart (Orders by Month)
plt.figure(figsize=(12,6))
sns.lineplot(data=dfs["orders_by_month"], x="month", y="total", marker="o", linewidth=2.5)
plt.title("Orders by Month", fontsize=16)
plt.xlabel("Month")
plt.ylabel("Orders")
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig("charts/orders_by_month_line.png")
plt.clf()

# 5. Histogram (Order Values)
plt.figure(figsize=(10,6))
sns.histplot(dfs["order_values"]["order_value"], bins=50, kde=True, color="purple")
plt.title("Distribution of Order Values", fontsize=16)
plt.xlabel("Order Value")
plt.ylabel("Frequency")
plt.tight_layout()
plt.savefig("charts/order_values_hist.png")
plt.clf()

# 6. Scatter plot (Items vs Order Value)
plt.figure(figsize=(10,6))
sns.scatterplot(data=dfs["order_items_vs_value"], x="items_count", y="order_value", alpha=0.6, color="darkred")
plt.title("Items Count vs Order Value", fontsize=16)
plt.xlabel("Items per Order")
plt.ylabel("Order Value")
plt.tight_layout()
plt.savefig("charts/items_vs_value_scatter.png")
plt.clf()

# ------------------------------
# 2. Интерактивные графики (Plotly)
# ------------------------------

# Линейный график
fig = px.line(
    dfs["orders_by_month"],
    x="month",
    y="total",
    title="Orders by Month (Interactive)"
)
fig.write_html("charts/orders_by_month_interactive.html")

# интерактивный с ползунком времени
query_slider = """
SELECT 
    DATE_TRUNC('month', o.order_purchase_timestamp) AS month,
    c.customer_state,
    COUNT(o.order_id) AS total_orders
FROM orders o
JOIN customers c ON o.customer_id = c.customer_id
GROUP BY month, c.customer_state;
"""

df_slider = pd.read_sql(query_slider, conn)

# все уникальные месяцы и штаты
all_months = df_slider["month"].unique()
all_states = df_slider["customer_state"].unique()

# создаём полный набор комбинаций
full_index = pd.MultiIndex.from_product([all_months, all_states], names=["month", "customer_state"])
df_full = pd.DataFrame(index=full_index).reset_index()

# объединяем с оригинальными данными
df_slider_full = pd.merge(
    df_full,
    df_slider,
    on=["month", "customer_state"],
    how="left"
).fillna({"total_orders": 0})

# сортируем
df_slider_full = df_slider_full.sort_values(["month", "total_orders"], ascending=[True, False])

# строим график
fig_slider = px.bar(
    df_slider_full,
    x="customer_state",
    y="total_orders",
    color="customer_state",
    animation_frame=df_slider_full["month"].dt.strftime("%Y-%m"),
    range_y=[0, df_slider_full["total_orders"].max()+100],
    title="Orders by State Over Time (Interactive with Slider)"
)

fig_slider.write_html("charts/orders_by_state_slider.html")

# ------------------------------
# 3. Экспорт в Excel с форматированием
# ------------------------------
excel_path = "exports/analytics_report.xlsx"
with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    for name, df in dfs.items():
        df.to_excel(writer, sheet_name=name, index=False)

wb = load_workbook(excel_path)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

wb.save(excel_path)

print(" Все задания выполнены! Графики в /charts, отчёт в /exports")
